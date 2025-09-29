# coding  : utf8
# Date   : 2022/7/28 - 11:04
# Author : Ding Ning

import os
import openpyxl
from openpyxl import load_workbook, Workbook
from library.file_operation import FileOperation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class ExcelOperation:
    def __init__(self, file_name):
        self.fo = FileOperation()
        self.log = self.fo.log
        self.file_name = file_name
        if not os.path.exists(file_name):
            ExcelOperation.create_excel_file(file_name)
        self.wb = load_workbook(file_name)
        # 获取sheet页
        self.sheet_name_list = self.wb.sheetnames

    @staticmethod
    def create_excel_file(file_name, *sheet_name):
        """
        新建Excel文件
        :param file_name:       文件路径+文件名
        :param sheet_name:      sheet页
        :return:
        """
        wb = openpyxl.Workbook()
        # 新建sheet页
        if sheet_name and len(sheet_name) > 1:
            for s in range(len(sheet_name)):
                if s == 0:
                    for i in wb:
                        i.title = sheet_name[0]
                else:
                    wb.create_sheet(str(sheet_name[s]))
        elif len(sheet_name) == 1:
            for i in wb:
                i.title = sheet_name[0]
        # 创建文件
        wb.save(file_name)

    @staticmethod
    def insert_sheet(file_name, sheet_name, sheet_index=-1):
        """
        新增sheet页
        :param file_name:
        :param sheet_name:
        :param sheet_index:     默认插在最后
        :return:
        """
        wb = load_workbook(file_name)
        sheet_name_list = wb.sheetnames
        if sheet_index == -1:
            sheet_index = len(sheet_name_list)
        else:
            sheet_index = sheet_index
        wb = openpyxl.load_workbook(file_name)
        wb.create_sheet(title=sheet_name, index=sheet_index)
        wb.save(file_name)

    def input_in_excel(self, input_content, x_loc=0, y_loc=0, sheet=0, save_status=False):
        """
        excel中输入值
        :param input_content:   输入内容，输入列表则在末尾添加一行
        :param x_loc:           x
        :param y_loc:           y
        :param sheet:           页数
        :param save_status:     是否自动保存
        :return:
        """
        # 加载sheet页
        sheet_name = self.sheet_name_list[sheet]
        ws = self.wb[sheet_name]
        # 判断表格是否为空
        # cell_data = self.get_data_from_excel(1, 1)
        if isinstance(input_content, list):
            # 添加
            # try:
            ws.append(input_content)
            # except Exception as e:
            #     self.log.info(e)
            #     self.log.info(input_content)
        else:
            # 输入单个值
            ws.cell(x_loc, y_loc).value = input_content
            # noinspection PyBroadException
            try:
                alpha = self.num_to_alpha(x_loc)
                cell_data = f'{alpha}{y_loc}'
                self.set_cel_border(cell_data)
            except Exception:
                pass
        if save_status:
            self.wb.save(self.file_name)

    def get_data_from_excel(self, x_loc, y_loc, sheet=0):
        """
        从excel中获取值
        :param x_loc:           x
        :param y_loc:           y
        :param sheet:           页数
        :return:
        """
        # 加载sheet页
        sheet_name = self.sheet_name_list[sheet]
        ws = self.wb[sheet_name]
        # 获取值
        cel = ws.cell(x_loc, y_loc)
        cel_value = cel.value
        return cel_value

    @staticmethod
    def num_to_alpha(num):
        """
        数字转换成字母
        :param num:         数字
        :return:
        """
        if 24 < num or num < 1:
            raise Exception('参数异常！')
        num += 64
        a = chr(num)
        return a

    def get_max_row_and_line(self, sheet=0):
        """
        获取sheet页的最大行列
        :param sheet:           页数
        :return:
        """
        # 加载sheet页
        sheet_name = self.sheet_name_list[sheet]
        ws = self.wb[sheet_name]
        max_row = ws.max_row
        max_line = ws.max_column
        return max_row, max_line

    def get_list_data_from_sheet(self, sheet=0):
        """
        按行读取数据
        :param sheet:           页数
        :return:
        """
        # 加载sheet页
        sheet_name = self.sheet_name_list[sheet]
        ws = self.wb[sheet_name]
        row_data = ws.rows
        rows = list()
        for row in list(row_data):
            part_l = list()
            for c in row:  # 把每行的每个单元格的值取出来，存放到case里
                part_l.append(c.value)
            rows.append(part_l)
        return rows

    def get_line_data(self, line_no, sheet=0):
        """
        获取指定行得数据
        :param line_no:
        :param sheet:
        :return:
        """
        line_data = []
        sheet_name = self.sheet_name_list[sheet]
        ws = self.wb[sheet_name]
        for i in ws[line_no]:
            line_data.append(i.value)
        return line_data

    def set_cel_border(self, cell, border_style='thin', color='000000', sheet=0, save_status=False):
        """
        设置边框
        :param cell:            单元格位置
        :param border_style:
        :param color:
        :param sheet:           页数
        :param save_status:     是否自动保存
        :return:
        """
        # 加载sheet页
        sheet_name = self.sheet_name_list[sheet]
        ws = self.wb[sheet_name]
        thin = Side(border_style=border_style, color=color)
        # 边框的位置
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws[cell].border = border
        if save_status:
            self.wb.save(self.file_name)

    def set_cel_font(self, cell, size=14, name='微软雅⿊', color='000000', sheet=0, save_status=False):
        """
        设置字体
        :param cell:            单元格位置
        :param size:
        :param name:
        :param color:
        :param sheet:           页数
        :param save_status:     是否自动保存
        :return:
        """
        # 加载sheet页
        sheet_name = self.sheet_name_list[sheet]
        ws = self.wb[sheet_name]
        font = Font(size=size, bold=True, name=name, color=color)
        ws[cell].font = font
        if save_status:
            self.wb.save(self.file_name)

    def set_cel_align(self, cell, horizontal='center', vertical='center', sheet=0, save_status=False):
        """
        对齐方式
        :param cell:            单元格位置
        :param horizontal:      ⽔平⽅向 center, left, right
        :param vertical:        垂直⽅向 center, top, bottom
        :param sheet:           页数
        :param save_status:     是否自动保存
        :return:
        """
        # 加载sheet页
        sheet_name = self.sheet_name_list[sheet]
        ws = self.wb[sheet_name]
        alight = Alignment(
            horizontal=horizontal,
            vertical=vertical
        )
        # 边框的位置
        ws[cell].alignment = alight
        if save_status:
            self.wb.save(self.file_name)

    def set_cel_fill(self, cell, fill_type='solid', start_color='FFFFFF', sheet=0, save_status=False):
        """
        填充颜色
        :param cell:
        :param fill_type:
        :param start_color:
        :param save_status:     是否自动保存
        :param sheet:           页数
        :return:
        """
        # 加载sheet页
        sheet_name = self.sheet_name_list[sheet]
        ws = self.wb[sheet_name]
        fill = PatternFill(fill_type=fill_type, start_color=start_color)
        # 边框的位置
        ws[cell].fill = fill
        if save_status:
            self.wb.save(self.file_name)

    def set_cel_wid_or_high(self, cell, cel_type='wid', num=20, sheet=0, save_status=False):
        """
        设置单元格宽度和高度
        :param cell:            单元格
        :param cel_type:        宽度
        :param num:             高度
        :param sheet:           页数
        :param save_status:     是否自动保存
        :return:
        """
        # 加载sheet页
        sheet_name = self.sheet_name_list[sheet]
        ws = self.wb[sheet_name]
        if cel_type == 'wid':
            ws.column_dimensions[cell].width = num
        else:
            ws.row_dimensions[cell].height = num
        self.wb.save(self.file_name)
        if save_status:
            self.wb.save(self.file_name)

    def merge_cells(self, range_string, start_column=0, end_row=0, end_column=0, merge_type='normal', sheet=0, umerge='merge'):
        """
        合并单元格
        :param range_string:        'A1:B3'/5
        :param start_column:        4
        :param end_row:             8
        :param end_column:          8
        :param merge_type:          合并方式
        :param sheet:               sheet页
        :param umerge:              合并或取消合并
        :return:
        """
        sheet_name = self.sheet_name_list[sheet]
        ws = self.wb[sheet_name]
        if merge_type == 'normal':
            if umerge == 'merge':
                ws.merge_cells(range_string=range_string)
            else:
                ws.unmerge_cells(range_string=range_string)
        else:
            if umerge == 'merge':
                ws.merge_cells(start_row=range_string,
                               start_column=start_column,
                               end_row=end_row,
                               end_column=end_column)
            else:
                ws.unmerge_cells(start_row=range_string,
                                 start_column=start_column,
                                 end_row=end_row,
                                 end_column=end_column)

    def disassemble_excel(self, dis_file_line):
        """
        将一个excel拆分成若干个文件
        :param dis_file_line:
        :return:
        """
        file_path, file_name = os.path.split(self.file_name)
        # 读取原始Excel文件
        df = self.get_list_data_from_sheet()
        excel_title = df.pop(0)
        # 拆分数据
        total_list = self.fo.divide_chunks(df, dis_file_line)
        total_list = list(total_list)
        for t in total_list:
            t.insert(0, excel_title)
        # 创建10个新的Excel文件
        for i, data_part in enumerate(total_list):
            output_file = f'{file_path}\\output_{i + 1}.xlsx'
            wb = Workbook()
            sheet = wb.active
            for d in data_part:
                sheet.append(d)
            wb.save(output_file)
        self.log.info('文件拆分结束！')

    def save_excel(self):
        """
        保存文件
        :return:
        """
        self.wb.save(self.file_name)


if __name__ == '__main__':
    file = r'C:\Users\Administrator\Desktop\AKUS\111\修罗\阿修罗-藏品批量销毁模板.xlsx'
    eo = ExcelOperation(file)
    # eo.get_data_from_excel(1, 1)
    # eo.set_cel_align('A1')
    # eo.input_in_excel(['pack文件路径', 'pack文件名', '开始时间戳', '开始时间', '结束时间戳', '结束时间', '测试结果'])
    # res = eo.num_to_alpha(2)
    # print(res)
    # ExcelOperation.insert_sheet(r'D:\Download\数据回传\3\trigger_test_res.xlsx', 'wqfqwfqwf')
    eo.disassemble_excel(120)
    # eo.save_excel()
    # res = eo.get_data_from_excel(2, 19, 1)
    # print(res)
    # eo.get_list_data_from_sheet()
